import os.path, sys
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
#from openpyxl.compat import range as xls_range
from openpyxl.styles import Border, Side, colors, Font
#from datetime import date
from . import hosp_config as config 
from poly.utils.files import get_name_tail

def extract(row):
    if row.nap_num is None:
        return None
    d = [ '' for i in range(8)]
    fam = row.fam or ''
    im = row.im or ' '
    ot = row.ot or ' '
    d[0] = '%s %s %s' % (row.fam, im[0].upper(), ot[0].upper())  
    d[1] = row.date_birth.strftime('%d.%m.%Y')
    d[2] = row.ds
    d[3] = row.specfic or ''
    d[4] = '%s' % row.nap_num
    d[5] = row.nap_date.strftime('%d.%m.%Y')
    
    if int(row.for_pom_hospl) > 0:
        d[6] = "ЭКСТР"
    
    if row.usl_ok > 0:
        d[7] = 'ДНЕВНОЙ СТ'
    
    return d
    
def make_report(app, year, month, procClass):
    
    sh1 = 'ЕИР'

    file = config.FILE_R
    _data = config.GET_HOSP
   
    mnt = int(month)
    base_dir = os.path.join( app.config['UPLOAD_FOLDER'],  config.BASE_DIR )
    xlr = os.path.join( base_dir, config.TPL_DIR, (file + '.xlsx'))
    xlw = os.path.join(
        base_dir, config.REPORT_DIR , (
            file + f'_0{mnt}_{get_name_tail(5)}.xlsx'
        )
    )
    #year = date.today().isocalendar()[0]
    period = 'За %s %s года' % (config.MONTHS[mnt-1], year)
    
    wb = load_workbook(filename = xlr)
    wb.active
    sheet = wb[sh1]
    sheet['D2'].value = period
    mo_border = Border(
        #left=Side(border_style='thin', color=colors.BLACK),
        #right=Side(border_style='thin', color=colors.BLACK),
        #top=Side(border_style='thin', color=colors.BLACK),
        bottom=Side(border_style='thin', color=colors.BLACK)
    )    
    mo_font = Font(
        name='Arial',
        size=16,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000')
    total_font = Font(name='Arial', size=13, bold=True)
    
    #qonn = psycopg2.connect("dbname=prive user=postgres password=boruh")
    qonn = app.config.db()
    qurs = qonn.cursor(cursor_factory=psycopg2.extras.NamedTupleCursor)

    
    # begin from 6 string
    rowXls = 6
    rcTotal = 0
    #print(procClass.mo)
    #qurs.execute(config.GET_MO)
    #for _mo in qurs.fetchall():
    for mo_hash, mo_name in procClass.mo.items():
        #print(f'{mo_hash} -- {mo_name}')
        if not bool(mo_name):
            #print(f'mo -- {mo_name}')
            continue
        mo_cell = sheet.cell(column=1, row = rowXls, value = mo_name)
        mo_cell.font=mo_font
        mo_cell.border= mo_border
        rowXls += 2
        qurs.execute(config.GET_HOSP, (mo_hash,))
        total_mo = 0
        for _hosp in qurs.fetchall():
           
            data = extract(_hosp)
            if data is None:
                continue
            for xcol, val in enumerate(data):
                sheet.cell(column=xcol+1, row=rowXls, value=val )
            rowXls += 1
            total_mo += 1
        c1 = sheet.cell(column=1, row=rowXls, value = 'Итого по МО')
        c2 = sheet.cell(column=2, row=rowXls, value = '%i' % total_mo)
        c1.font = total_font
        c2.font = total_font
        rowXls += 2
        #print('-- %s --' % rcTotal, end='\r')
        rcTotal += total_mo
    
    rowXls += 2
    c1 = sheet.cell(column=1, row=rowXls, value = 'Итого по в ФАЙЛЕ')
    c2 = sheet.cell(column=2, row=rowXls, value = '%i' % rcTotal)
    c1.font = total_font
    c2.font = total_font   
    
    wb.save(xlw)
    wb.close()
    qurs.close()
    qonn.close()
    
    return xlw