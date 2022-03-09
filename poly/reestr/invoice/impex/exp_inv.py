
import os, types
from datetime import date
from pathlib import Path
import psycopg2
import psycopg2.extras
from flask import g
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors
from poly.utils.files import get_name_tail
from poly.reestr.invoice.impex import config

sn = types.SimpleNamespace()

def get_mo_smo_name(sn, mo_code, smo, cfg):
    assert hasattr(sn, 'qurs'), 'get_mo_smo_name:: Нет связи с БД'

    _code= mo_code[-3:] # just last 3 digits i.e short code
    sn.qurs.execute(cfg.GET_MO_NAME, ( _code, ) )
    mq= sn.qurs.fetchone()
    if mq:
        mo_name= mq[0]
    else:
        mo_name= cfg.STUB_MO
    if smo > 0:
        #ins= 25000 + int(smo)
        sn.qurs.execute(cfg.GET_SMO_NAME, ( smo, ))
        mq= sn.qurs.fetchone()
        if mq:
            smo_name=  mq[0]
        else:
            smo_name= cfg.STUB_SMO
    else:
        smo_name= ''

    return mo_name, smo_name


def extract(row):
    d = [ '' for i in range(20)]
    d[0] = row.nhistory
    fam = row.fam or ' '
    im = row.im or ' '
    ot = row.ot or ' '
    d[1] = '%s %s %s' % (row.fam, im[0].upper(), ot[0].upper())
    d[2] = row.w
    d[3] = row.dr
    #d[4] = ''
    #docser = row.docser or ' '
    #docnum = row.docnum or ' '
    #d[5] = '%s %s' % (docser, docnum)
    #d[6], d[7], d[8] = '', '', ''
    spolis = row.spolis or ''
    npolis = row.npolis or ''
    d[9] = '%s %s' % (spolis, npolis)
    #''.join(i for i in row.p_num if i.isdigit())
    # re.sub("\D", "", )
    d[10] = row.vidpom
    d[11] = row.ds1
    d[12] = '%s~%s' % (row.date_z_1, row.date_z_2)
    d[13] = 1
    d[14] = row.profil
    d[15] = row.prvs
    d[16] = price = row.sumv
    #d[17] = row.foms_price
    if row.sank_it:
        #price -= row.sank_it
        if price:
            d[5]= 'МЭК'
        else:
            d[5]= 'ХЭК'
        price= 0.00
    d[17] = price

    d[18] = row.rslt
    #d[19] = row.ishod

    return d


def for_foms(dex, rc):
    # d - list from extract
    d = ['' for i in range(20)]

    d[0] = rc

    # nusl, fio
    d[1], d[2] = dex[0], dex[1]

    # pol
    #d[3] = ['м', 'ж'].index( dex[2] ) - 1
    d[3]= dex[2]

    # date_birth, place_birth
    d[4], d[5] = dex[3], dex[4]

    # d[6] # document
    # d[7] # snils

    # polis
    d[8] = dex[9]

    # vid_pom
    d[9] = dex[10]

    # DS
    d[10] = dex[11]

    # date
    d[11], d[12] = dex[12].split("~")

    i = 13
    for v in dex [13:]:
        d[ i ] = v
        i += 1

    return d


def data_source_init(db, is_calc):
    global sn
    sn.qurs = db.cursor(cursor_factory=psycopg2.extras.NamedTupleCursor)
    if len(is_calc) == 0:
        _data = config.COUNT_INV
    else:
        _data= config.COUNT_MO
    sn.qurs.execute(_data)
    rc= sn.qurs.fetchone()
    if bool(rc[0]):
        return True
    sn.qurs.close()
    return False

def data_source_get(is_calc):
    global sn
    if len(is_calc) == 0:
        _data = config.GET_ROW_INV
    else:
        _data= config.GET_ROW_MO
    sn.qurs.execute(_data)
    return sn.qurs.fetchall()

def data_source_close():
    global sn
    sn.qurs.close()


def exp_inv(
        app: object, # current app object
        db: object, # db connection
        mo_code: str, # long MO_CODE i.e 250796
        smo: int, # int (0,  25011, 25016 )
        month: str, #str(2) 01..12
        year: str, #str(4) 2020
        typ: int, # 1-5
        inv_path: str, # path to the data folder
        is_calc='' # flag str if not empty then self calculated reestr
    ): # -> (int, str):

    global sn

    if not data_source_init(db, is_calc):
        return (0, '')

    m= int(month)
    tpl= config.TYPE[typ-1][2]

    sh1 = 'Лист1'

    xtpl = f'{tpl}.xlsx'
    xlr = os.path.join(inv_path, 'tpl', xtpl)

    xout = f'{tpl}{is_calc}_{smo}_{month}_{year}_{get_name_tail(5)}.xlsx'
    xlw = os.path.join(inv_path, xout)

    period = 'За %s %s года' % (app.config['MONTH'][m-1], year)

    wb = load_workbook(filename = xlr)
    wb.active
    sheet = wb[sh1]

    mo_name, smo_name = get_mo_smo_name(sn, mo_code, smo, config)

    if smo == 0:
        sheet['C4'].value = period
        # begin from 17 string
        cntRowXls = 17
    else:
        # local config has MOS dict with mo_code as KEY and tuple(OGRN, ) as value
        sheet['E2'].value = '%s ОГРН %s' % (mo_name, app.config['MOS'][mo_code][0])
        sheet['E7'].value = smo_name
        sheet['J1'].value = period
        # begin from 13 string
        cntRowXls = 13

    border = Border(
        left=Side(border_style='thin', color=colors.BLACK),
        right=Side(border_style='thin', color=colors.BLACK),
        top=Side(border_style='thin', color=colors.BLACK),
        bottom=Side(border_style='thin', color=colors.BLACK)
    )

    rcTotal = 1
    irang=20 # 20 cells in row
    #irang=21 # 21 cells in row
    dc= 0

    for row in data_source_get(is_calc):

        data = extract(row)
        if smo == 0:
            data = for_foms(data, rcTotal)

        for xrow in range(cntRowXls, cntRowXls+1):
            for xcol in range(1, irang):
                try:
                    c = sheet.cell(column=xcol, row=xrow, value= data[ xcol-1 ])
                    c.border = border
                except Exception as e:
                    app.logger.debug('\n row %s ' % rcTotal)
                    app.logger.debug(data)
                    raise e
            cntRowXls += 1
            rcTotal += 1

    wb.save(xlw)
    wb.close()
    data_source_close()

    return (rcTotal-1, xlw)
