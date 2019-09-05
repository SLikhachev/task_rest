
# makes by paraclin code calculating file

import os
from collections import defaultdict
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.styles import Border, Side, colors
from poly.reestr.invoice.impex.exp_inv import get_mo_smo_name
from poly.reestr.invoice.impex import config as imp_cfg
from poly.reestr.invoice.tarif import config as tarif_config
from poly.reestr.invoice.calc import config

def calc_pmu(app: object, insurer: str, month: str, yar: str, typ: int, inv_path: str) -> (int, str):
    
    qonn= app.config.db()
    qurs = qonn.cursor(cursor_factory=psycopg2.extras.NamedTupleCursor)
    
    tpl= imp_cfg.TYPE[typ-1][2]
    sh1 = 'Лист1'
    xtpl = f'{tpl}.xlsx'
    xlr = os.path.join(inv_path, 'tpl', xtpl)   
    xout = f'{tpl}_0{insurer}_{month}-{yar}.xlsx'
    
    mon= int(month)
    smo= int(insurer)
    
    xlw = os.path.join(inv_path, xout)   
    year= f'20{yar}'
    period = 'За %s %s года' % (app.config['MONTH'][mon-1], year)
    
    mo_name, smo_name = get_mo_smo_name(app, qurs, insurer, imp_cfg)
    
    wb = load_workbook(filename = xlr)
    wb.active
    sheet = wb[sh1]
    sheet['B15'].value = '%s   %s' % (period, smo_name)

    border = Border(
        left=Side(border_style='thin', color=colors.BLACK),
        right=Side(border_style='thin', color=colors.BLACK),
        top=Side(border_style='thin', color=colors.BLACK),
        bottom=Side(border_style='thin', color=colors.BLACK)
    )
    
    get_talon = config.GET_TALON
    get_code = config.GET_CODE

    codes = defaultdict(int)
    qurs.execute(get_talon, (mon, smo) )
    rows= qurs.fetchall()
    for row in rows:
        qurs.execute(get_code, ( row.tal_num, ))
        tal_codes= qurs.fetchall()
        for c in tal_codes:
             codes[c.code_usl] += int(c.kol_usl)

    if len( codes.keys() ) == 0:
        wb.close()
        qurs.close()
        qonn.close()
        return 0, ''
        
    get_tarif= config.GET_TARIF % tarif_config.VZAIMO_PMU_TABLE[1]
    qurs.execute(get_tarif)
    tarif = {}
    for el in qurs.fetchall():
       tarif[ el.code ] = (el.tarif, el.name)

    cnt = 21
    rc = 1
    for code in sorted(codes.keys()):
    #for code, kol in codes.items():
        price= 0.0
        _name= 'Услуга не тарифицирована'
        usl = tarif.get(code, None)
        if usl is not None:
            price= usl[0]
            _name= usl[1]
        
        #data = [ code, name, price, kol, kol*price ]
        kol = codes[code]
        data = ( code, _name,  price, kol, kol*price)
        for xrow in range(cnt, cnt+1):
            for xcol in range(1, 6):
                c = sheet.cell(column=xcol, row=xrow, value= data[ xcol-1 ])
                c.border = border
        cnt += 1  
        rc += 1
    
    for xcol in range(1, 6):
        val=''
        if xcol==1:
            val="ИТОГО"
        if xcol==4:
            d = f'D{cnt-1}'
            val=f'=SUM(D21:D{cnt-1})'
        if xcol==5:
            d = f'E{cnt-1}'
            val=f'=SUM(E21:E{cnt-1})'
        c = sheet.cell(column=xcol, row=cnt, value= val)
        c.border = border
    
    sheet.cell(column=2, row=cnt+4, value= "Генеральный директор")
    sheet.cell(column=3, row=cnt+4, value= "ДЕНИСОВА С. А.")
    sheet.cell(column=2, row=cnt+6, value= "Исполнитель   202-51-45   Чамбайшин Ю.А.")	

    wb.save(xlw)
    wb.close()
    qurs.close()
    qonn.close()

    return rc, xlw