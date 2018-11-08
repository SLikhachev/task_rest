import os
import sys
from openpyxl import load_workbook
from openpyxl.compat import range as xls_range
from openpyxl.styles import Border, Side, colors, Font
import psycopg2
import psycopg2.extras
from poly.report.common.volum import volum_config as config


def make_report(app, db, month, year):

    mnt = int(month)
    file = config.FILE_R
    base_dir = os.path.join(app.config['UPLOAD_FOLDER'], config.BASE_DIR)
    xlr = os.path.join(base_dir, config.TPL_DIR, (file + '.xlsx'))
    xlw = os.path.join(base_dir, config.REPORT_DIR, (file + '_0%s_%s.xlsx' % (mnt, year)))
    #app.logger.debug('%s' % xlw.replace('\\', '+'))
    # year = date.today().isocalendar()[0]
    period = ' За %s %s года' % (app.config['MONTH'][mnt - 1], year)

    wb = load_workbook(filename=xlr)
    #sheet = wb.active
    ws = wb['Объемы']

    smo_border = Border(
        # left=Side(border_style='thin', color=colors.BLACK),
        # right=Side(border_style='thin', color=colors.BLACK),
        # top=Side(border_style='thin', color=colors.BLACK),
        bottom=Side(border_style='thin', color=colors.BLACK)
    )
    smo_font = Font(
        name='Arial',
        size=13,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000')
    total_font = Font(name='Arial', size=14, bold=True)
    c = ws['F2']
    c.value = period
    c.font = total_font

    #qur = db.cursor(cursor_factory=psycopg2.extras.NamedTupleCursor)
    qur = db.cursor()
    qur.execute(config.THIS_MONTH, (year, month))
    row = 6 # start row in template
    data_idx = list( range(config._ROW['pol_ambul_visits'], config._ROW['travma_ambul_persons'] + 1))
    offset = 0
    for r in qur.fetchall():
        insurer = r[config._ROW['insurer'] - offset]
        if config.INSURERS.get(insurer, None) is not None:
            hdr = config.INSURERS[insurer]
        else:
            hdr = 'Неизвестный страховщик ' + insurer
        hdr += ' за месяц ' + app.config['MONTH'][mnt - 1]
        hdr_cell = ws.cell(row=row, column=1, value=hdr)
        hdr_cell.font = smo_font
        #hdr_cell.border = smo_border
        row += 1

        for col_idx, val_idx in enumerate(data_idx):
            c = ws.cell(row = row, column=col_idx+1, value= r[val_idx-offset])
            c.border = smo_border
            c.font = smo_font
        row += 2
    row += 3

    qur.execute(config.THIS_YEAR, (year,))
    offset = config.TOTAL_OFFSET
    for r in qur.fetchall():
        insurer = r[ config._ROW['insurer'] - offset]
        if config.INSURERS.get(insurer, None) is not None:
             hdr = config.INSURERS[ insurer ]
        else:
            hdr = 'Неизвестный страховщик ' + insurer
        hdr +=  ' с начала %s года' % year
        hdr_cell = ws.cell(row=row, column=1, value=hdr)
        hdr_cell.font = smo_font
        # hdr_cell.border = smo_border
        row += 1

        for col_idx, val_idx in enumerate( data_idx ):
            c = ws.cell(row=row, column=col_idx + 1, value=r[val_idx - offset])
            c.border = smo_border
            c.font = smo_font
        row += 2

    wb.save(xlw)
    wb.close()

    qur.close()
    return xlw

