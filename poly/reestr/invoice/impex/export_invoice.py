""" definition of the calss for export data to the XLSX file """

import os
from typing import NamedTuple
from decimal import getcontext, Decimal
from psycopg2 import sql as psy_sql
from flask import g
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors
from poly.utils.files import get_name_tail
from poly.reestr.invoice.impex import config


class SqlExport:
    """ base class for export """

    def __init__(self,
        flask_app: object, sql: object, mo_code: str, smo: int,
        month: str, year: str, typ: int, export_folder: str, is_calc: str=''):
        """
           @param: flask_app: object - current flask app object
           @param: sql: object - Sql provider context manager
           @param: mo_code: str - full MO_CODE i.e '250799'
           @param: smo: int, code of (0,  25011, 25016 )
           @param: month: str(2) - 01..12
           @param: year: str(4) - '2020'
           @param: typ: int, - 1-5 package type
           @param: export_folder: str, path to the folder where file will be stored
           @param: is_calc='': str - if not empty then self calculated reestr
            used for substring in output file name
        """
        self.app = flask_app
        self.sql = sql

        # clean same values
        assert hasattr(sql, 'qurs'), "SqlExport: Нет связи с БД"

        self.qurs = sql.qurs
        self.mo_code = mo_code
        self.smo = smo
        self.month = month
        self.year = year
        self.typ = typ
        self.export_dir = export_folder
        self.calc = is_calc
        self.workbook = None
        self.sheet_title = 'Лист1'
        self.total_sum = Decimal(0.00)
        self.total_sum_column=-1
        # icode for select SQL statement "A05%", "A06%"
        self.icode = ''

    def check_table_name(self, table_name):
        assert hasattr(self.sql, table_name), f"Экспорт реестра: имя таблицы: {table_name} БД не определено"

    def get_mo_smo_name(self):
        """ select and return (SMO, MO) names from DB """
        _code= self.mo_code[-3:] # last 3 digits i.e short code
        self.qurs.execute(config.GET_MO_NAME, ( _code, ) )
        _mo= self.qurs.fetchone()
        if _mo:
            mo_name= _mo[0]
        else:
            mo_name= config.STUB_MO
        if self.smo > 0:
            #ins= 25000 + int(smo)
            self.qurs.execute(config.GET_SMO_NAME, ( self.smo, ))
            smo= self.qurs.fetchone()
            if smo:
                smo_name=  smo[0]
            else:
                smo_name= config.STUB_SMO
        else:
            smo_name= ''

        return mo_name, smo_name

    def border(self):
        return Border(
            left=Side(border_style='thin', color=colors.BLACK),
            right=Side(border_style='thin', color=colors.BLACK),
            top=Side(border_style='thin', color=colors.BLACK),
            bottom=Side(border_style='thin', color=colors.BLACK)
        )

    def init_workbook(self, sheet_title=None):
        """
        Init workbook once

        If the workbook has already been initialized, return it immediately.
        Otherwise, load the template workbook and set up the output file name and path.
        """
        if self.workbook is not None:
            return self.workbook

        # Get the template name from the type config

        self._tpl_name = config.TYPE[self.typ-1][2]

        # Set up the template file name and path
        self.tpl_name = f'{self._tpl_name}.xlsx'
        self.tpl_abspath = os.path.join(self.export_dir, 'tpl', self.tpl_name)

        # Check if the template file exists
        if not os.path.exists(self.tpl_abspath):
            raise AttributeError(f"Шаблон {self.tpl_abspath} не найден")

        # Set up the output file name and path
        self.xlsout_fname = \
            f'{self._tpl_name}{self.calc}_{self.smo}_{self.month}_{self.year}_{get_name_tail(5)}.xlsx'

        self.xlsout_abspath = os.path.join(self.export_dir, self.xlsout_fname)

        # Set up the period string
        self.period = f"За {self.app.config['MONTH'][int(self.month)-1]} {self.year} года"

        # Load the template workbook
        self.workbook = load_workbook(filename = self.tpl_abspath)

        # Set the active sheet
        self.workbook.active

        return self.workbook

    def set_sent(self, talon_row: NamedTuple):
        pass

    def close_workbook(self, rows, close_workbook: bool=True):
        assert self.workbook is not None, "Workbook is not initialized"
        self.workbook.save(self.xlsout_abspath)
        if close_workbook:
            self.workbook.close()
            self.workbook = None
        return rows, self.xlsout_fname


class SqlExportInvoice(SqlExport):
    """ export invoice DB table to xlsx file """

    def __init__(self,
        flask_app: object, sql: object, mo_code: str, smo: int,
        month: str, year: str, typ: int, export_folder: str, is_calc=''):

        super().__init__(flask_app, sql, mo_code, smo,
            month, year, typ, export_folder, is_calc)

    def extract_for_smo(self, row: NamedTuple, cells_in_row) -> list:
        """ make the xlsx's row list for the SMO invoice's reestr"""
        getcontext().prec = 2
        prop = lambda attr: getattr(row, attr, '')
        sprop = lambda attr: getattr(row, attr, '  ')
        fst = lambda attr: sprop(attr)[0].upper() if sprop(attr) else ''
        _d = [ '' for _ in range(cells_in_row)]
        _d[0] = prop('nhistory')
        _d[1] = f"{sprop('fam')} {fst('im')}. {fst('ot')}."
        _d[2] = prop('w')
        _d[3] = prop('dr')
        #d[4] = ''
        #docser = row.docser or ' '
        #docnum = row.docnum or ' '
        #d[5] = '%s %s' % (docser, docnum)
        #d[6], d[7], d[8] = '', '', ''
        _d[9] = f"{prop('spolis')} {prop('npolis') or prop('enp')}"
        #''.join(i for i in row.p_num if i.isdigit())
        # re.sub("\D", "", )
        _d[10] = prop('vidpom')
        _d[11] = prop('ds1')
        _d[12] = f"{prop('date_z_1')}~{prop('date_z_2')}"
        _d[13] = 1
        _d[14] = prop('profi')
        _d[15] = prop('prvs')
        _d[16] = price = Decimal(prop('sump'))
        #d[17] = row.foms_price
        if prop('sump') == 0.00:
            #price -= row.sank_it
            if price:
                _d[5]= 'МЭК'
            else:
                _d[5]= 'ХЭК'
            price= 0.00
        _d[17] = price

        _d[18] = prop('rslt')
        #d[19] = row.ishod

        self.total_sum += Decimal(price)
        return _d


    def extarct_for_foms(self, smo_row: list, record_num: int, cells_in_row: int):
        """ make xlsx's row list for TFOMS invoice's reestr"""
        # smo_row - list from extract_for_smo
        _d = ['' for _ in range(cells_in_row)]

        _d[0] = record_num

        # idcase, fio
        _d[1], _d[2] = smo_row[0], smo_row[1]

        # pol
        _d[3]= smo_row[2]

        # date_birth, place_birth
        _d[4], _d[5] = smo_row[3], smo_row[4]

        # d[6] # document
        # d[7] # snils

        # polis
        _d[8] = smo_row[9]

        # vid_pom
        _d[9] = smo_row[10]

        # DS
        _d[10] = smo_row[11]

        # date
        _d[11], _d[12] = smo_row[12].split("~")

        idx = 13
        for _v in smo_row[13:]:
            _d[ idx ] = _v
            idx += 1

        return _d


    def check_tables_exists(self, table_name: str=''):
        """ check DB tables for export exists"""

        self.check_table_name('inv_table')
        _data = psy_sql.SQL(config.COUNT_INV_TMP).format(self.sql.inv_table)
        if len(self.calc) > 0:
            pass
            #_data= config.COUNT_MO
        self.qurs.execute(_data)
        return True if self.qurs.fetchone() else False

    def select_export_data(self):
        """ select all from DB """
        _data = psy_sql.SQL(config.GET_ROW_INV_TMP).format(self.sql.inv_table)
        if len(self.calc) > 0:
            pass
            #_data= config.GET_ROW_MO
        self.qurs.execute(_data)
        return self.qurs.fetchall()

    def prepare_to_foms(self, sheet):
        sheet['C4'].value = self.period
        #self.total_sum_column=20
        # begin from 17th string
        return 17, 20

    def prepare_to_smo(self, sheet):
        mo_name, smo_name = self.get_mo_smo_name()
        # ======= code for SMO ==========
        # local config has MOS dict with mo_code as KEY and tuple(OGRN, ) as value
        sheet['E2'].value = f"{mo_name} ОГРН {self.app.config['MOS'][self.mo_code][0]}"
        sheet['E7'].value = smo_name
        sheet['J1'].value = self.period
        self.total_sum_column=20
        # begin from 13th string
        return 13, 20

    def prepare_sheet(self, sheet):
        if self.smo == 0:
            return self.prepare_to_foms(sheet)
        else:
            return self.prepare_to_smo(sheet)

    def extract_data(self, row, cells_in_row):
        data = self.extract_for_smo(row, cells_in_row)
        if self.smo == 0:
            data = self.extarct_for_foms(data, self.rc_total, cells_in_row)
        return data

    def export(self,
        sheet_title: str="Лист1",
        sent: bool=False,
        close_workbook: bool=True
    ) -> tuple:
        """
        Main function to export data to xlsx file.

        Args:
            sheet_title (str): title of the sheet
            sent (bool): mark as sent if used
            close_workbook (bool): close workbook after export

        Returns:
            tuple: result of the export operation
        """
        if not self.check_tables_exists():
            return (-1, "Нет экспотрной таблицы БД")

        wb = self.init_workbook()
        sheet = wb[sheet_title]

        current_row, cells_in_row = self.prepare_sheet(sheet)

        self.total_sum = Decimal(0.00)
        border = self.border()
        rc_total = 1
        for row in self.select_export_data():
            try:
                data = self.extract_data(row, cells_in_row)
            except Exception as _ex:
                #print(_ex)
                #print(row)
                raise _ex
            for xrow in range(current_row, current_row+1):
                for xcol in range(1, cells_in_row+1): # +1 because of range is [1..n) interval
                    try:
                        cell = sheet.cell(column=xcol, row=xrow, value= data[ xcol-1 ])
                        cell.border = border
                        if type(cell.value) == Decimal:
                            cell.number_format = '### ### ##0.00'
                    except Exception as exc:
                        self.app.logger.debug(f"\n row {rc_total} {exc}")
                        self.app.logger.debug(data)
                        raise exc
                current_row += 1
                rc_total += 1

            ## mark as sent if used
            if sent:
                self.set_sent(row)

        ## insert sum total
        if self.total_sum_column > 0:
            #print(f"Total sum: {self.total_sum}")
            sheet.cell(
                column=self.total_sum_column-1, # prev columnt
                row=current_row+2, value= "ИТОГО:",
            )
            cell = sheet.cell(
                column=self.total_sum_column,
                row=current_row+2, value=self.total_sum,
            )
            cell.number_format = '# ### ##0.00'

        return self.close_workbook(rc_total-1, close_workbook=close_workbook)
